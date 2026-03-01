# -*- coding: utf-8 -*-
"""
hd_classify6.py – AI classifier + storytelling (sheet "Storytelling")
-------------------------------------------------------------------
Usage examples:
  python3 hd_classify6.py \
    --input report.xlsx \
    --output report_AI_v6.xlsx \
    --provider openai \
    --model gpt-4o-mini \
    --subject-col "Problém" \
    --desc-col "Popis problému" \
    --rpm 30 \
    --story

Notes:
- Requires OPENAI_API_KEY or AZURE_OPENAI_API_KEY/ENDPOINT/DEPLOYMENT env vars.
- Writes classification columns and, when --story is set and output is .xlsx, a new sheet "Storytelling"
  with a concise narrative summary for management.
"""
import os, re, json, time, argparse, logging
from typing import Dict, List, Tuple
import pandas as pd
import requests
from datetime import datetime

# -----------------------------
# Categories / keywords (copied from v5)
# -----------------------------
CATEGORIES = [
    "Instalace / aktualizace software",
    "Údržba / čištění disků a souborů",
    "Databáze / Oracle / MCC / Virtuan",
    "Interní aplikace (Unique, NemExpress, Argus…)",
    "Tiskárny a skenery",
    "Síť / Wi-Fi / VPN / GlobalProtect",
    "Office / Outlook / Teams",
    "Hardware (notebook, monitor, periferie)",
    "Přístupy / oprávnění / účty",
    "Cloud / OneDrive / SharePoint",
    "Mobil / iPhone / Android",
    "Ostatní",
]

KEYWORDS: Dict[str, List[str]] = {
    "Instalace / aktualizace software": [
        "instalace", "doinstalace", "nainstalovat", "naistalovat",
        "aktualizace", "update", "upgrade", "nová verze",
        "setup", "install", "reinstall", "re-instalace",
        "patch", "aplikace", "program"
    ],
    "Údržba / čištění disků a souborů": [
        "úklid", "uklid", "čištění", "cisteni",
        "disk c", "disk d", "volné místo", "vymazat",
        "mazání", "smazat", "odstranit soubory"
    ],
    "Databáze / Oracle / MCC / Virtuan": [
        "oracle", "mcc", "virtuan", "elfa", "db",
        "databáze", "database", "chyba 500", "sql",
        "patch oracle"
    ],
    "Interní aplikace (Unique, NemExpress, Argus…)": [
        "unique", "nemexpress", "argus", "mcc", "elfa",
        "interní aplikace", "internal app"
    ],
    "Tiskárny a skenery": [
        "tisk", "tisknout", "tiskárna", "tiskarna",
        "scanner", "scan", "skener", "mxm",
        "toner", "print"
    ],
    "Síť / Wi-Fi / VPN / GlobalProtect": [
        "wifi", "wi-fi", "síť", "sit",
        "vpn", "gp", "global protect", "připojení",
        "nejde se připojit", "network", "disconnect"
    ],
    "Office / Outlook / Teams": [
        "outlook", "office", "teams", "email",
        "mail", "e-mail", "schránka", "calendar",
        "meeting", "teams meeting", "pošta"
    ],
    "Hardware (notebook, monitor, periferie)": [
        "notebook", "laptop", "monitor", "obrazovka",
        "hardware", "ventilátor", "usb", "klávesnice",
        "myš", "charger", "baterie", "battery"
    ],
    "Přístupy / oprávnění / účty": [
        "přístup", "oprávnění", "pristup",
        "account", "heslo", "password", "login",
        "přihlášení", "nelze se přihlásit",
        "rights", "permissions"
    ],
    "Cloud / OneDrive / SharePoint": [
        "onedrive", "sharepoint", "cloud",
        "upload", "sdílená složka", "shared folder"
    ],
    "Mobil / iPhone / Android": [
        "iphone", "android", "mobil",
        "sim", "volání", "call", "roaming"
    ],
    "Ostatní": []
}

ALIASES: List[Tuple[str, str]] = [
    ("global connect", "Síť / Wi-Fi / VPN / GlobalProtect"),
    ("gp", "Síť / Wi-Fi / VPN / GlobalProtect"),
    ("nelze tisknout", "Tiskárny a skenery"),
    ("nemexpress", "Interní aplikace (Unique, NemExpress, Argus…)"),
    ("unique", "Interní aplikace (Unique, NemExpress, Argus…)"),
    ("argus", "Interní aplikace (Unique, NemExpress, Argus…)"),
    ("oracle", "Databáze / Oracle / MCC / Virtuan"),
    ("mcc", "Databáze / Oracle / MCC / Virtuan"),
]


def normalize(t: str) -> str:
    return re.sub(r"\s+", " ", str(t or "")).strip().lower()


def category_by_keywords(text: str) -> Tuple[str, int, str]:
    """Rule engine – returns (cat, conf, explanation)"""
    txt = normalize(text)
    # Aliases – strong signal
    for phrase, cat in ALIASES:
        if phrase in txt:
            return cat, 95, f"alias:{phrase}"
    # Best keyword hit
    best_cat = None
    best_hits = 0
    for cat, words in KEYWORDS.items():
        hits = sum(1 for w in words if w in txt)
        if hits > best_hits:
            best_hits = hits
            best_cat = cat
    if best_hits >= 3:
        return best_cat, 90, "keywords-strong"
    if best_hits == 2:
        return best_cat, 75, "keywords-medium"
    if best_hits == 1:
        return best_cat, 55, "keywords-weak"
    return None, 0, "no-match"


def validate_category(cat: str) -> str:
    if not cat:
        return "Ostatní"
    for c in CATEGORIES:
        if cat.lower() == c.lower():
            return c
    cat_low = cat.lower()
    for c in CATEGORIES:
        if any(word in cat_low for word in c.lower().split()):
            return c
    return "Ostatní"


class LLMClient:
    def __init__(self, provider: str, model: str, rpm: int):
        self.provider = provider
        self.model = model
        self.rpm = rpm
        self.api_key = None
        if provider == "openai":
            # FIX: jen jedna varianta proměnné – bez zpětných lomítek
            self.api_key = os.getenv("OPENAI_API_KEY")
            self.url = "https://api.openai.com/v1/chat/completions"
            self.header_name = "Authorization"
            self.header_value = f"Bearer {self.api_key}"
        else:
            # Azure OpenAI
            self.api_key = os.getenv("AZURE_OPENAI_API_KEY")
            endpoint = os.getenv("AZURE_OPENAI_ENDPOINT")
            deployment = os.getenv("AZURE_OPENAI_DEPLOYMENT")
            version = os.getenv("AZURE_OPENAI_API_VERSION", "2024-02-15-preview")
            self.url = f"{endpoint}/openai/deployments/{deployment}/chat/completions?api-version={version}"
            self.header_name = "api-key"
            self.header_value = self.api_key
        if not self.api_key:
            raise ValueError("Chybí API key!")
        self._last = 0
        self.min_interval = 60 / max(1, rpm)

    def _rate_limit(self):
        now = time.time()
        w = (self._last + self.min_interval) - now
        if w > 0:
            time.sleep(w)
        self._last = time.time()

    def _extract_json(self, content: str) -> Dict:
        m = re.search(r"\{.*\}", content, flags=re.S)
        if not m:
            return {"category": "Ostatní", "confidence": 0, "explanation": "no-json"}
        text = m.group(0)
        try:
            data = json.loads(text)
        except Exception as e:
            return {"category": "Ostatní", "confidence": 0, "explanation": f"json-error:{e}"}
        data["category"] = validate_category(data.get("category"))
        try:
            data["confidence"] = int(data.get("confidence", 0))
        except Exception:
            data["confidence"] = 0
        return data

    def ask(self, text: str, strict: bool = False) -> Dict:
        system = (
            "Jsi zkušený IT analytik. Urči jednu kategorii."
            if not strict
            else "Jsi zkušený IT analytik. Vyber jednu nejbližší kategorii. 'Ostatní' použij jen v úplné nouzi."
        )
        # FIX: souvislý f-string, žádné osamocené { text } mimo řetězec
        user_prompt = f"""
Text tiketu:
\"\"\"
{text}
\"\"\"
Vrať JSON:
{{
  "category": "...",
  "confidence": 0-100,
  "explanation": "..."
}}
"""
        payload = {
            "model": self.model,
            "temperature": 0,
            "messages": [
                {"role": "system", "content": system},
                {"role": "user", "content": user_prompt}
            ]
        }
        self._rate_limit()
        r = requests.post(self.url, headers={self.header_name: self.header_value}, json=payload, timeout=60)
        if r.status_code >= 400:
            raise RuntimeError(f"HTTP {r.status_code}: {r.text}")
        content = r.json()["choices"][0]["message"]["content"]
        return self._extract_json(content)

    def ask_story(self, structure: Dict) -> str:
        """Generate a Czech storytelling summary for a management sheet."""
        system = (
            "Jsi seniorní IT service manager a analytik. Vytváříš stručné, věcné a akční shrnutí incidentů. "
            "Piš česky, profesionálně, bez marketingu."
        )
        user_prompt = (
            "Vytvoř maximálně 10 odrážek + krátké Executive summary (3-5 vět) z těchto agregovaných dat."
            " Zaměř se na trendy, opakující se témata, SLA/MTTR, a akční doporučení (max 5), včetně kandidátů na automatizaci."
            " Zůstaň konkrétní, cituj čísla. Nepiš tabulky ani kód."
            f"\n\nDATA:\n{json.dumps(structure, ensure_ascii=False, indent=2)}"
        )
        payload = {
            "model": self.model,
            "temperature": 0.2,
            "messages": [
                {"role": "system", "content": system},
                {"role": "user", "content": user_prompt}
            ]
        }
        self._rate_limit()
        r = requests.post(self.url, headers={self.header_name: self.header_value}, json=payload, timeout=90)
        if r.status_code >= 400:
            raise RuntimeError(f"HTTP {r.status_code}: {r.text}")
        return r.json()["choices"][0]["message"]["content"].strip()


def load_file(path: str) -> pd.DataFrame:
    if path.lower().endswith((".xlsx", ".xls")):
        return pd.read_excel(path, engine="openpyxl")
    return pd.read_csv(path, encoding="utf-8-sig")


def compute_aggregates(df: pd.DataFrame, story_cols: Dict) -> Dict:
    """Prepare a compact structure used for story generation."""
    total = len(df)
    if "Stav" in df.columns:
        resolved_mask = df["Stav"].astype(str).str.lower().str.contains("vyřešeno|vyreseno|resolved")
    else:
        resolved_mask = pd.Series([False] * total)
    resolved = int(resolved_mask.sum())
    open_cnt = int(total - resolved)

    mttr_vals = []
    if "Vykázaný čas" in df.columns:
        for v in df["Vykázaný čas"]:
            try:
                mt = float(v)
                if mt > 0:
                    mttr_vals.append(mt)
            except Exception:
                pass
    mttr = round(sum(mttr_vals) / len(mttr_vals), 1) if mttr_vals else None
    median = round(float(pd.Series(mttr_vals).median()), 1) if mttr_vals else None

    # top categories if present
    cat_col = story_cols.get("ai_category_col", "AI_Category")
    top_cats = (df[cat_col].value_counts().head(8).to_dict() if cat_col in df.columns else {})

    # top normalized subjects (based on "Problém")
    subj_col = story_cols.get("subject_col")
    top_subjects = {}
    if subj_col in df.columns:
        norm = df[subj_col].astype(str).str.strip().str.lower()
        top_subjects = norm.value_counts().head(10).to_dict()

    # departments if present
    by_dept = df["Oddělení"].value_counts().to_dict() if "Oddělení" in df.columns else {}

    last_month = None
    if "Vytvořeno" in df.columns:
        try:
            # Excel serials or datetimes
            ser = pd.to_datetime(df["Vytvořeno"], errors="coerce", unit="d", origin="1899-12-30")
            if ser.notna().any():
                last = ser.max()
                last_month = last.strftime("%Y-%m")
        except Exception:
            try:
                ser = pd.to_datetime(df["Vytvořeno"], errors="coerce")
                if ser.notna().any():
                    last_month = ser.max().strftime("%Y-%m")
            except Exception:
                pass

    structure = {
        "period_last_month": last_month,
        "total_tickets": total,
        "resolved": resolved,
        "open": open_cnt,
        "mttr_min": mttr,
        "median_min": median,
        "top_ai_categories": top_cats,
        "top_subjects_norm": top_subjects,
        "tickets_by_department": by_dept,
    }
    return structure


def write_story_sheet(xlsx_path: str, story_text: str, structure: Dict):
    """Append/replace sheet 'Storytelling' in the XLSX with text lines and a compact metadata table."""
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet

    wb = load_workbook(xlsx_path)
    if "Storytelling" in wb.sheetnames:
        ws = wb["Storytelling"]
        wb.remove(ws)
    ws: Worksheet = wb.create_sheet("Storytelling")

    # Title + text
    ws["A1"] = "AI shrnutí"
    ws["A2"] = story_text

    # Metadata block (simple key-value list)
    row = 4
    ws[f"A{row}"] = "Meta"; row += 1
    for k, v in structure.items():
        ws[f"A{row}"] = str(k)
        ws[f"B{row}"] = json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else ("" if v is None else v)
        row += 1

    wb.save(xlsx_path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--output", required=True)
    ap.add_argument("--subject-col", default="Problém")
    ap.add_argument("--desc-col", default="Popis problému")
    ap.add_argument("--provider", choices=["openai", "azure"], required=True)
    ap.add_argument("--model", default="gpt-4o-mini")
    ap.add_argument("--rpm", type=int, default=60)
    ap.add_argument("--max-rows", type=int, default=None)
    ap.add_argument("--offline-only", action="store_true")
    ap.add_argument("--story", action="store_true", help="Vytvořit sheet 'Storytelling' se shrnutím")
    args = ap.parse_args()

    logging.basicConfig(level=logging.INFO)
    df = load_file(args.input)
    if args.max_rows:
        df = df.head(args.max_rows)
    if args.subject_col not in df.columns or args.desc_col not in df.columns:
        raise ValueError("Chybí sloupce Problém / Popis problému.")

    llm = None
    if not args.offline_only:
        llm = LLMClient(args.provider, args.model, rpm=args.rpm)

    out_cat, out_conf, out_exp, out_method = [], [], [], []
    for idx, row in df.iterrows():
        text = f"Predmet: {row.get(args.subject_col,'')}\nPopis: {row.get(args.desc_col,'')}"
        # RULE PASS
        cat0, conf0, exp0 = category_by_keywords(text)
        best_cat = cat0
        best_conf = conf0
        best_exp = exp0
        method = "rules"
        # LLM PASS
        if (conf0 < 80) and llm:
            r1 = llm.ask(text, strict=False)
            cat1, conf1 = r1["category"], r1["confidence"]
            if conf1 > best_conf:
                best_cat, best_conf, best_exp = cat1, conf1, r1.get("explanation", "")
                method = "llm-pass1"
        # STRICT PASS
        if best_cat == "Ostatní" and llm:
            r2 = llm.ask(text, strict=True)
            if r2["category"] != "Ostatní" and r2["confidence"] >= best_conf:
                best_cat, best_conf, best_exp = r2["category"], r2["confidence"], r2["explanation"]
                method = "llm-pass2"
        out_cat.append(validate_category(best_cat))
        out_conf.append(best_conf)
        out_exp.append(best_exp)
        out_method.append(method)
        if (idx + 1) % 50 == 0:
            logging.info(f"Zpracováno {idx+1} řádků…")

    df["AI_Category"] = out_cat
    df["AI_Confidence"] = out_conf
    df["AI_Explanation"] = out_exp
    df["AI_Method"] = out_method

    # Save main output
    if args.output.lower().endswith(".xlsx"):
        df.to_excel(args.output, index=False, engine="openpyxl")
    else:
        df.to_csv(args.output, index=False, encoding="utf-8-sig")

    # Storytelling – only if requested and xlsx output
    if args.story and args.output.lower().endswith(".xlsx"):
        story_cols = {"ai_category_col": "AI_Category", "subject_col": args.subject_col}
        structure = compute_aggregates(df, story_cols)
        if llm:
            story_text = llm.ask_story(structure)
        else:
            story_text = (
                "(Offline režim) Storytelling je vypnutý – není dostupný LLM klient. "
                "Pro generování použij --provider openai/azure a přidej --story."
            )
        write_story_sheet(args.output, story_text, structure)

    print("\n=== HOTOVO (v6) ===")
    print(f"Výstup: {args.output}")


if __name__ == "__main__":
    main()
